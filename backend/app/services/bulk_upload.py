"""
Bulk Upload Service
Handles Excel template generation and parsing for bulk submissions
"""

import io
import logging
from typing import List, Dict, Any, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

logger = logging.getLogger(__name__)

# Template column definitions matching SubmissionCreate schema
TEMPLATE_COLUMNS = [
    {"key": "candidate_name", "header": "Candidate Name *", "width": 25, "example": "John Doe"},
    {"key": "candidate_email", "header": "Email *", "width": 30, "example": "john@example.com"},
    {"key": "github_url", "header": "GitHub URL *", "width": 50, "example": "https://github.com/username/repo"},
    {"key": "hosted_url", "header": "Hosted URL (optional)", "width": 45, "example": "https://myapp.vercel.app"},
    {"key": "video_url", "header": "Video URL (optional)", "width": 45, "example": "https://drive.google.com/..."},
]


class BulkUploadService:
    """Service for handling bulk submission uploads via Excel"""

    def generate_template(self) -> bytes:
        """
        Generate Excel template with headers and example row.

        Returns:
            bytes: Excel file content
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Bulk Submissions"

        # Header styling - Cyan background with black text
        header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Example row styling - Gray italic text
        example_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        example_font = Font(italic=True, color="666666", size=10)

        # Border style
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Write headers
        for col_idx, col_def in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = header_alignment
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = col_def["width"]

        # Write example row (row 2)
        for col_idx, col_def in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_def["example"])
            cell.fill = example_fill
            cell.font = example_font
            cell.border = thin_border

        # Add instruction row (row 4)
        ws.cell(row=4, column=1, value="Instructions:")
        ws.cell(row=4, column=1).font = Font(bold=True)

        instructions = [
            "1. Fill in candidate details starting from row 2 (or delete the example row)",
            "2. Columns marked with * are required",
            "3. GitHub URL must be a valid repository URL",
            "4. Hosted URL and Video URL are optional",
            "5. Delete this instruction section before uploading"
        ]

        for idx, instruction in enumerate(instructions, 5):
            ws.cell(row=idx, column=1, value=instruction)
            ws.cell(row=idx, column=1).font = Font(color="666666", size=9)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def parse_excel(self, file_content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Parse uploaded Excel file and return list of submissions.

        Args:
            file_content: Raw Excel file bytes

        Returns:
            Tuple of (valid_submissions, errors)
        """
        try:
            wb = load_workbook(io.BytesIO(file_content))
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            return [], [{"row": 0, "error": f"Invalid Excel file: {str(e)}", "data": {}}]

        ws = wb.active
        if ws is None:
            return [], [{"row": 0, "error": "No active sheet found", "data": {}}]

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")

        # Map headers to keys using flexible matching
        header_to_key = {}
        for col_idx, header in enumerate(headers):
            header_lower = header.lower().replace(" ", "_").replace("*", "").replace("(", "").replace(")", "")
            for col_def in TEMPLATE_COLUMNS:
                if col_def["key"] in header_lower or header_lower in col_def["key"]:
                    header_to_key[col_idx] = col_def["key"]
                    break

        submissions = []
        errors = []

        # Process data rows (skip header row)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Skip empty rows
            if not any(row):
                continue

            # Skip instruction rows (contain specific text)
            first_cell = str(row[0]) if row[0] else ""
            if "instruction" in first_cell.lower() or first_cell.startswith(("1.", "2.", "3.", "4.", "5.")):
                continue

            submission = {}
            for col_idx, value in enumerate(row):
                if col_idx in header_to_key:
                    key = header_to_key[col_idx]
                    submission[key] = str(value).strip() if value else None

            # Validate required fields
            missing = []
            for col_def in TEMPLATE_COLUMNS:
                if "(optional)" not in col_def["header"]:
                    if not submission.get(col_def["key"]):
                        missing.append(col_def["header"].replace(" *", ""))

            if missing:
                errors.append({
                    "row": row_idx,
                    "error": f"Missing required fields: {', '.join(missing)}",
                    "data": submission
                })
            else:
                # Validate email format
                email = submission.get("candidate_email", "")
                if email and "@" not in email:
                    errors.append({
                        "row": row_idx,
                        "error": f"Invalid email format: {email}",
                        "data": submission
                    })
                    continue

                # Validate GitHub URL
                github_url = submission.get("github_url", "")
                if github_url and "github.com" not in github_url:
                    errors.append({
                        "row": row_idx,
                        "error": f"Invalid GitHub URL: {github_url}",
                        "data": submission
                    })
                    continue

                submissions.append(submission)

        logger.info(f"Parsed Excel: {len(submissions)} valid submissions, {len(errors)} errors")
        return submissions, errors


# Singleton instance
bulk_upload_service = BulkUploadService()
